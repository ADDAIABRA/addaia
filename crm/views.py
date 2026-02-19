"""
Views do CRM - Coleta e visualização de leads.
"""
import csv
import io
import threading
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from pagamentos.decoradores_acesso import exigir_assinante_ativo

from .models import Coleta, Lead
from .forms import ColetarLeadsForm
from .services.places_collector import run_coleta


@exigir_assinante_ativo
def crm(request):
    """Dashboard CRM com links para Coletar e Ver Leads."""
    titulo_pagina = 'CRM'
    descricao_pagina = 'Bem-vindo ao CRM da plataforma.'
    context = {
        'titulo_pagina': titulo_pagina,
        'descricao_pagina': descricao_pagina
    }
    return render(
        request=request,
        template_name='app/crm/crm.html',
        context=context
    )


@exigir_assinante_ativo
def coletar_leads(request):
    """Formulário para filtrar e iniciar coleta. POST cria Coleta e redireciona."""
    acesso = getattr(request.user, 'acesso', None)
    if not acesso:
        messages.error(request, "Você precisa de um plano ativo para coletar leads.")
        return redirect('planos')

    limite = acesso.leads_limite_mensal or 0
    consumidos = acesso.leads_consumidos_mes or 0
    disponiveis = max(0, limite - consumidos)

    if limite <= 0:
        messages.warning(request, "Seu plano não inclui coleta de leads.")
        return redirect('planos')

    if not acesso.tem_leads_disponiveis():
        messages.warning(request, "Você atingiu o limite de leads deste mês.")
        return redirect('ver_leads')

    if request.method == 'POST':
        form = ColetarLeadsForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            coleta = Coleta.objects.create(
                usuario=request.user,
                keyword=cd['keyword'],
                cidade=cd['cidade'],
                bairro=cd.get('bairro') or '',
                usar_raio=cd.get('usar_raio', False),
                raio_km=cd.get('raio_km') if cd.get('usar_raio') else None,
                status='em_andamento'
            )
            thread = threading.Thread(target=run_coleta, args=(coleta.id,))
            thread.daemon = True
            thread.start()
            return redirect(f"{reverse('ver_leads')}?coleta={coleta.id}")
        else:
            messages.error(request, "Corrija os erros no formulário.")
    else:
        form = ColetarLeadsForm()

    context = {
        'titulo_pagina': 'Coletar Leads',
        'descricao_pagina': 'Defina os critérios e inicie a coleta.',
        'form': form,
        'leads_disponiveis': disponiveis,
        'leads_limite': limite,
    }
    return render(request, 'app/crm/coletar_leads.html', context)


@exigir_assinante_ativo
def ver_leads(request):
    """Lista de leads do usuário. Se ?coleta=N, mostra modal e polling."""
    coleta = None
    coleta_id = request.GET.get('coleta')
    if coleta_id:
        try:
            coleta = Coleta.objects.get(id=int(coleta_id), usuario=request.user)
        except (ValueError, Coleta.DoesNotExist):
            pass

    leads = Lead.objects.filter(usuario=request.user).select_related('coleta').order_by('-criado_em')[:500]

    context = {
        'titulo_pagina': 'Meus Leads',
        'descricao_pagina': 'Leads coletados para sua prospecção.',
        'leads': leads,
        'coleta': coleta,
    }
    return render(request, 'app/crm/ver_leads.html', context)


@exigir_assinante_ativo
@require_GET
def leads_stream(request, coleta_id):
    """
    Endpoint JSON para polling. Retorna leads da coleta desde since_id.
    Apenas se coleta pertence ao usuário.
    """
    coleta = get_object_or_404(Coleta, id=coleta_id, usuario=request.user)
    since_id = request.GET.get('since_id', '0')
    try:
        since_id = int(since_id)
    except ValueError:
        since_id = 0

    leads_qs = Lead.objects.filter(coleta=coleta, id__gt=since_id).order_by('id')
    leads = [
        {
            'id': l.id,
            'categoria': l.categoria or '',
            'cidade': l.cidade or '',
            'bairro': l.bairro or '',
            'nome': l.nome,
            'telefone': l.telefone or '',
            'endereco': l.endereco or '',
            'site': l.site or '',
            'nota': str(l.nota) if l.nota is not None else '',
            'total_avaliacoes': l.total_avaliacoes,
        }
        for l in leads_qs
    ]

    return JsonResponse({
        'leads': leads,
        'coleta_status': coleta.status,
        'total': Lead.objects.filter(coleta=coleta).count(),
    })


def _get_leads_queryset(user):
    """Retorna queryset de leads do usuário para exportação."""
    return Lead.objects.filter(usuario=user).select_related('coleta').order_by('-criado_em')


def _lead_row(lead):
    """Retorna lista de valores de um lead para exportação."""
    return [
        lead.categoria or '',
        lead.cidade or '',
        lead.bairro or '',
        lead.nome,
        lead.telefone or '',
        lead.endereco or '',
        lead.site or '',
        str(lead.nota) if lead.nota is not None else '',
        lead.total_avaliacoes,
    ]


@exigir_assinante_ativo
@require_GET
def export_leads_csv(request):
    """Exporta leads do usuário em CSV."""
    leads = _get_leads_queryset(request.user)
    header = ['Categoria', 'Cidade', 'Bairro', 'Nome', 'Telefone', 'Endereço', 'Site', 'Nota', 'Avaliações']
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="leads.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(header)
    for lead in leads:
        writer.writerow(_lead_row(lead))
    return response


@exigir_assinante_ativo
@require_GET
def export_leads_excel(request):
    """Exporta leads do usuário em Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        messages.error(request, "Biblioteca openpyxl não instalada. Execute: pip install openpyxl")
        return redirect('ver_leads')

    leads = _get_leads_queryset(request.user)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leads'

    header = ['Categoria', 'Cidade', 'Bairro', 'Nome', 'Telefone', 'Endereço', 'Site', 'Nota', 'Avaliações']
    for col, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for row_idx, lead in enumerate(leads, 2):
        for col_idx, val in enumerate(_lead_row(lead), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="leads.xlsx"'
    return response
